import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import os

def export_to_excel(tree, file_name_prefix="export"):
    """
    Xuất dữ liệu từ Treeview ra file Excel với định dạng đẹp mắt.
    
    Args:
        tree: Đối tượng Treeview chứa dữ liệu.
        file_name_prefix: Tiền tố tên file Excel (mặc định là "export").
    """
    try:
        # Lấy tiêu đề từ Treeview
        columns = [tree.heading(col)["text"] for col in tree["columns"]]
        
        # Lấy dữ liệu từ Treeview
        data = []
        for item in tree.get_children():
            row = tree.item(item)["values"]
            data.append(row)

        # Tạo workbook và worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Export"

        # Định dạng tiêu đề
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Ghi tiêu đề vào file Excel
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Ghi dữ liệu vào file Excel
        for row_idx, row in enumerate(data, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

        # Định dạng viền cho tất cả các ô
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Tô màu xen kẽ cho các dòng
        even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Định dạng các ô
        for row in ws.iter_rows(min_row=1, max_row=len(data) + 1, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = thin_border
                # Tô màu xen kẽ cho dòng chẵn
                if cell.row % 2 == 0 and cell.row != 1:  # Bỏ qua tiêu đề
                    cell.fill = even_row_fill
                # Căn giữa cho tất cả các ô (trừ tiêu đề đã căn giữa ở trên)
                if cell.row != 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Định dạng số cho các cột lương (nếu có)
        number_columns = ["Lương cơ bản", "Lương theo giờ", "Tiền tăng ca", "Tổng lương"]
        for col_idx, col_name in enumerate(columns, 1):
            if col_name in number_columns:
                for row_idx in range(2, len(data) + 2):  # Bắt đầu từ dòng 2 (dòng dữ liệu)
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        cell.value = float(str(cell.value).replace(",", ""))  # Chuyển chuỗi có dấu "," thành số
                        cell.number_format = "#,##0"  # Định dạng số với dấu phân cách hàng nghìn
                    except (ValueError, TypeError):
                        pass  # Nếu không phải số thì bỏ qua

        # Tự động điều chỉnh độ rộng cột
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for cell in column:
                try:
                    # Đo độ dài nội dung ô
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
                except:
                    pass
            # Điều chỉnh độ rộng cột (thêm padding 2 để không bị sát chữ)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Mở hộp thoại để người dùng chọn nơi lưu file
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{file_name_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            title="Lưu file Excel"
        )

        if file_path:
            # Lưu file Excel
            wb.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất dữ liệu ra file Excel: {os.path.basename(file_path)}")
        else:
            messagebox.showwarning("Hủy", "Bạn đã hủy lưu file!")

    except Exception as e:
        messagebox.showerror("Lỗi", f"Xuất file Excel thất bại: {str(e)}")